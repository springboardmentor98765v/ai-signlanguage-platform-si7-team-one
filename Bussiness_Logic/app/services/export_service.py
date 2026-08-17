import csv
import io
from uuid import UUID
from datetime import datetime, timezone
from typing import List, Optional

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, HRFlowable, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT

from app.services.practice_service import practice_store
from app.services.assessment_service import assessment_store
from app.services.gamification_service import compute_gamification
from app.services.certification_service import certification_store
from app.schemas.practice import SessionStatus


def _grade_from_score(score: float) -> str:
    if score >= 90: return "A"
    if score >= 75: return "B"
    if score >= 60: return "C"
    if score >= 40: return "D"
    return "F"


def _header_row(ws, headers: list, fill_color: str = "1565C0"):
    try:
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise RuntimeError("openpyxl not installed")
    header_fill = PatternFill("solid", fgColor=fill_color)
    header_font = Font(color="FFFFFF", bold=True)
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = 20


def _pdf_doc(buffer, title: str, subtitle: str, learner_name: str):
    """Returns (doc, story_start, styles) ready to append rows to."""
    doc = SimpleDocTemplate(buffer, pagesize=A4,
        rightMargin=2*cm, leftMargin=2*cm, topMargin=2.5*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("t", parent=styles["Title"], fontSize=20,
        textColor=colors.HexColor("#1a237e"), alignment=TA_CENTER, spaceAfter=4)
    sub_style = ParagraphStyle("s", parent=styles["Normal"], fontSize=11,
        textColor=colors.HexColor("#37474f"), alignment=TA_CENTER, spaceAfter=12)
    story = [
        Paragraph(title, title_style),
        Paragraph(subtitle, sub_style),
        Paragraph(f"Learner: {learner_name}", sub_style),
        HRFlowable(width="100%", thickness=2, color=colors.HexColor("#1a237e"), spaceAfter=16),
    ]
    return doc, story, styles


# ── EXISTING EXPORTS (unchanged) ─────────────────────────────────────

def generate_learner_progress_csv(user_id: UUID) -> bytes:
    """One row per session: date, duration, accuracy, score, grade, signs practiced."""
    sessions = practice_store.get_sessions_by_user(user_id)
    sessions_sorted = sorted(sessions, key=lambda s: s.started_at)

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "Session Date", "Status", "Duration (seconds)", "Signs Practiced",
        "Correct Predictions", "Total Predictions", "Accuracy (%)", "Score", "Grade",
    ])

    for session in sessions_sorted:
        assessment = assessment_store.get(session.session_id)
        if assessment and assessment.total_predictions > 0:
            accuracy = round(assessment.correct_predictions / assessment.total_predictions * 100, 2)
            avg_score = round(assessment.score_sum / assessment.total_predictions, 2)
            grade = _grade_from_score(avg_score)
            signs = len(assessment.sign_stats)
            correct = assessment.correct_predictions
            total = assessment.total_predictions
        else:
            accuracy = avg_score = 0.0
            grade = "N/A"
            signs = correct = total = 0

        duration = 0
        if session.ended_at and session.started_at:
            duration = int((session.ended_at - session.started_at).total_seconds())

        writer.writerow([
            session.started_at.strftime("%Y-%m-%d %H:%M UTC"),
            session.status.value, duration, signs, correct, total, accuracy, avg_score, grade,
        ])

    return output.getvalue().encode("utf-8")


def generate_learner_progress_excel(user_id: UUID) -> bytes:
    """Excel version of the learner progress export."""
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError("openpyxl not installed")

    sessions = practice_store.get_sessions_by_user(user_id)
    sessions_sorted = sorted(sessions, key=lambda s: s.started_at)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Progress Report"
    _header_row(ws, ["Session Date", "Status", "Duration (s)", "Signs Practiced",
                      "Correct", "Total", "Accuracy (%)", "Score", "Grade"])

    for session in sessions_sorted:
        assessment = assessment_store.get(session.session_id)
        if assessment and assessment.total_predictions > 0:
            accuracy = round(assessment.correct_predictions / assessment.total_predictions * 100, 2)
            avg_score = round(assessment.score_sum / assessment.total_predictions, 2)
            grade = _grade_from_score(avg_score)
            signs = len(assessment.sign_stats)
            correct = assessment.correct_predictions
            total = assessment.total_predictions
        else:
            accuracy = avg_score = 0.0
            grade = "N/A"
            signs = correct = total = 0

        duration = 0
        if session.ended_at and session.started_at:
            duration = int((session.ended_at - session.started_at).total_seconds())

        ws.append([session.started_at.strftime("%Y-%m-%d %H:%M UTC"),
                   session.status.value, duration, signs, correct, total, accuracy, avg_score, grade])

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def generate_class_summary_csv() -> bytes:
    """One row per learner: user_id, total sessions, completed, avg accuracy, badges, weak signs."""
    all_sessions = list(practice_store._sessions.values())
    user_ids = list(set(s.user_id for s in all_sessions))

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["User ID", "Total Sessions", "Completed Sessions",
                     "Average Accuracy (%)", "Grade", "Badges Earned", "Weak Signs"])

    for user_id in user_ids:
        user_sessions = [s for s in all_sessions if s.user_id == user_id]
        completed = [s for s in user_sessions if s.status == SessionStatus.completed]
        accuracies, combined_sign_stats = [], {}

        for session in user_sessions:
            assessment = assessment_store.get(session.session_id)
            if assessment and assessment.total_predictions > 0:
                accuracies.append(assessment.correct_predictions / assessment.total_predictions * 100)
                for sign, stats in assessment.sign_stats.items():
                    entry = combined_sign_stats.setdefault(sign, {"correct": 0, "total": 0})
                    entry["correct"] += stats["correct"]
                    entry["total"] += stats["total"]

        avg_accuracy = round(sum(accuracies) / len(accuracies), 2) if accuracies else 0.0
        grade = _grade_from_score(avg_accuracy)
        weak_signs = [s for s, st in combined_sign_stats.items()
                      if st["total"] >= 2 and (st["correct"] / st["total"] * 100) < 60.0]
        gamification = compute_gamification(user_id)

        writer.writerow([str(user_id), len(user_sessions), len(completed), avg_accuracy,
                         grade, gamification.total_badges_earned,
                         ", ".join(weak_signs) if weak_signs else "None"])

    return output.getvalue().encode("utf-8")


def generate_class_summary_excel() -> bytes:
    """Excel version of the class summary export."""
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError("openpyxl not installed")

    all_sessions = list(practice_store._sessions.values())
    user_ids = list(set(s.user_id for s in all_sessions))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Class Summary"
    _header_row(ws, ["User ID", "Total Sessions", "Completed",
                      "Avg Accuracy (%)", "Grade", "Badges Earned", "Weak Signs"], "1A237E")

    for user_id in user_ids:
        user_sessions = [s for s in all_sessions if s.user_id == user_id]
        completed = [s for s in user_sessions if s.status == SessionStatus.completed]
        accuracies, combined_sign_stats = [], {}

        for session in user_sessions:
            assessment = assessment_store.get(session.session_id)
            if assessment and assessment.total_predictions > 0:
                accuracies.append(assessment.correct_predictions / assessment.total_predictions * 100)
                for sign, stats in assessment.sign_stats.items():
                    entry = combined_sign_stats.setdefault(sign, {"correct": 0, "total": 0})
                    entry["correct"] += stats["correct"]
                    entry["total"] += stats["total"]

        avg_accuracy = round(sum(accuracies) / len(accuracies), 2) if accuracies else 0.0
        grade = _grade_from_score(avg_accuracy)
        weak_signs = [s for s, st in combined_sign_stats.items()
                      if st["total"] >= 2 and (st["correct"] / st["total"] * 100) < 60.0]
        gamification = compute_gamification(user_id)

        ws.append([str(user_id), len(user_sessions), len(completed), avg_accuracy,
                   grade, gamification.total_badges_earned,
                   ", ".join(weak_signs) if weak_signs else "None"])

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# ── NEW REPORT TYPES (Milestone 4, Day 4) ────────────────────────────
# All 4 missing report types from the M4 SRS — each available as both
# PDF (ReportLab, same style as progress_service.py) and Excel (openpyxl).


# 1. LEARNING REPORT — signs practiced, time per session, session frequency

def generate_learning_report_pdf(user_id: UUID, learner_name: str) -> bytes:
    """Learning report: session frequency, signs practiced per session, total practice time."""
    sessions = sorted(practice_store.get_sessions_by_user(user_id), key=lambda s: s.started_at)
    buffer = io.BytesIO()
    doc, story, styles = _pdf_doc(buffer, "LEARNING REPORT",
                                   "AI-Powered Sign Language Learning Platform", learner_name)
    body = ParagraphStyle("b", parent=styles["Normal"], fontSize=11, spaceAfter=6)

    total_minutes = sum(
        int((s.ended_at - s.started_at).total_seconds()) // 60
        for s in sessions if s.ended_at and s.started_at
    )
    story.append(Paragraph(f"Total Sessions: <b>{len(sessions)}</b>", body))
    story.append(Paragraph(f"Total Practice Time: <b>{total_minutes} minutes</b>", body))
    story.append(Spacer(1, 0.5*cm))

    table_data = [["#", "Date", "Duration (min)", "Signs Practiced", "Status"]]
    for i, session in enumerate(sessions, 1):
        assessment = assessment_store.get(session.session_id)
        signs = len(assessment.sign_stats) if assessment else 0
        duration = 0
        if session.ended_at and session.started_at:
            duration = int((session.ended_at - session.started_at).total_seconds()) // 60
        table_data.append([
            str(i),
            session.started_at.strftime("%Y-%m-%d"),
            str(duration),
            str(signs),
            session.status.value,
        ])

    t = Table(table_data, colWidths=[1.2*cm, 4*cm, 3.5*cm, 3.5*cm, 3.5*cm])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1565c0")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e0e0e0")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f5f5f5")]),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
    ]))
    story.append(t)

    doc.build(story)
    pdf_bytes = buffer.getvalue()
    buffer.close()
    return pdf_bytes


def generate_learning_report_excel(user_id: UUID) -> bytes:
    """Excel version of the learning report."""
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError("openpyxl not installed")

    sessions = sorted(practice_store.get_sessions_by_user(user_id), key=lambda s: s.started_at)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Learning Report"
    _header_row(ws, ["#", "Date", "Duration (min)", "Signs Practiced", "Status"])

    for i, session in enumerate(sessions, 1):
        assessment = assessment_store.get(session.session_id)
        signs = len(assessment.sign_stats) if assessment else 0
        duration = 0
        if session.ended_at and session.started_at:
            duration = int((session.ended_at - session.started_at).total_seconds()) // 60
        ws.append([i, session.started_at.strftime("%Y-%m-%d"),
                   duration, signs, session.status.value])

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# 2. ASSESSMENT REPORT — per-session scores, grade breakdown

def generate_assessment_report_pdf(user_id: UUID, learner_name: str) -> bytes:
    """Assessment report: per-session weighted scores, accuracy, grade."""
    sessions = sorted(practice_store.get_sessions_by_user(user_id), key=lambda s: s.started_at)
    buffer = io.BytesIO()
    doc, story, styles = _pdf_doc(buffer, "ASSESSMENT REPORT",
                                   "AI-Powered Sign Language Learning Platform", learner_name)
    body = ParagraphStyle("b", parent=styles["Normal"], fontSize=11, spaceAfter=6)

    scored = [(s, assessment_store.get(s.session_id)) for s in sessions]
    scored = [(s, a) for s, a in scored if a and a.total_predictions > 0]

    if scored:
        all_scores = [a.score_sum / a.total_predictions for _, a in scored]
        overall_avg = round(sum(all_scores) / len(all_scores), 2)
        story.append(Paragraph(f"Overall Average Score: <b>{overall_avg}</b> "
                                f"(Grade: <b>{_grade_from_score(overall_avg)}</b>)", body))
        story.append(Spacer(1, 0.4*cm))

    table_data = [["#", "Date", "Total Attempts", "Correct", "Accuracy (%)", "Avg Score", "Grade"]]
    for i, (session, assessment) in enumerate(scored, 1):
        accuracy = round(assessment.correct_predictions / assessment.total_predictions * 100, 2)
        avg_score = round(assessment.score_sum / assessment.total_predictions, 2)
        table_data.append([
            str(i),
            session.started_at.strftime("%Y-%m-%d"),
            str(assessment.total_predictions),
            str(assessment.correct_predictions),
            str(accuracy),
            str(avg_score),
            _grade_from_score(avg_score),
        ])

    t = Table(table_data, colWidths=[1*cm, 3.5*cm, 3*cm, 2.5*cm, 3*cm, 3*cm, 2.5*cm])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a237e")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e0e0e0")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f5f5f5")]),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
    ]))
    story.append(t)

    doc.build(story)
    pdf_bytes = buffer.getvalue()
    buffer.close()
    return pdf_bytes


def generate_assessment_report_excel(user_id: UUID) -> bytes:
    """Excel version of the assessment report."""
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError("openpyxl not installed")

    sessions = sorted(practice_store.get_sessions_by_user(user_id), key=lambda s: s.started_at)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Assessment Report"
    _header_row(ws, ["#", "Date", "Total Attempts", "Correct", "Accuracy (%)", "Avg Score", "Grade"], "1a237e")

    for i, session in enumerate(sessions, 1):
        assessment = assessment_store.get(session.session_id)
        if not assessment or assessment.total_predictions == 0:
            continue
        accuracy = round(assessment.correct_predictions / assessment.total_predictions * 100, 2)
        avg_score = round(assessment.score_sum / assessment.total_predictions, 2)
        ws.append([i, session.started_at.strftime("%Y-%m-%d"),
                   assessment.total_predictions, assessment.correct_predictions,
                   accuracy, avg_score, _grade_from_score(avg_score)])

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# 3. ACCURACY REPORT — per-sign accuracy breakdown

def generate_accuracy_report_pdf(user_id: UUID, learner_name: str) -> bytes:
    """Accuracy report: per-sign correct/total/accuracy, sorted weakest first."""
    sessions = practice_store.get_sessions_by_user(user_id)
    combined: dict = {}
    for session in sessions:
        assessment = assessment_store.get(session.session_id)
        if not assessment:
            continue
        for sign, stats in assessment.sign_stats.items():
            entry = combined.setdefault(sign, {"correct": 0, "total": 0})
            entry["correct"] += stats["correct"]
            entry["total"] += stats["total"]

    # Sort weakest first
    sign_rows = sorted(
        [(sign, st["correct"], st["total"],
          round(st["correct"] / st["total"] * 100, 1) if st["total"] else 0.0)
         for sign, st in combined.items()],
        key=lambda x: x[3]
    )

    buffer = io.BytesIO()
    doc, story, styles = _pdf_doc(buffer, "ACCURACY REPORT",
                                   "AI-Powered Sign Language Learning Platform", learner_name)
    body = ParagraphStyle("b", parent=styles["Normal"], fontSize=11, spaceAfter=6)
    story.append(Paragraph(f"Signs Practiced: <b>{len(sign_rows)}</b>", body))
    story.append(Spacer(1, 0.4*cm))

    table_data = [["Sign", "Correct", "Total Attempts", "Accuracy (%)", "Status"]]
    for sign, correct, total, accuracy in sign_rows:
        status = "Weak" if accuracy < 60 else ("Strong" if accuracy >= 80 else "OK")
        table_data.append([sign, str(correct), str(total), str(accuracy), status])

    t = Table(table_data, colWidths=[3*cm, 3*cm, 4*cm, 4*cm, 3.5*cm])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2e7d32")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e0e0e0")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f5f5f5")]),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
    ]))
    story.append(t)

    doc.build(story)
    pdf_bytes = buffer.getvalue()
    buffer.close()
    return pdf_bytes


def generate_accuracy_report_excel(user_id: UUID) -> bytes:
    """Excel version of the accuracy report."""
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError("openpyxl not installed")

    sessions = practice_store.get_sessions_by_user(user_id)
    combined: dict = {}
    for session in sessions:
        assessment = assessment_store.get(session.session_id)
        if not assessment:
            continue
        for sign, stats in assessment.sign_stats.items():
            entry = combined.setdefault(sign, {"correct": 0, "total": 0})
            entry["correct"] += stats["correct"]
            entry["total"] += stats["total"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Accuracy Report"
    _header_row(ws, ["Sign", "Correct", "Total Attempts", "Accuracy (%)", "Status"], "2e7d32")

    for sign, st in sorted(combined.items(),
                            key=lambda x: (x[1]["correct"] / x[1]["total"]) if x[1]["total"] else 0):
        accuracy = round(st["correct"] / st["total"] * 100, 1) if st["total"] else 0.0
        status = "Weak" if accuracy < 60 else ("Strong" if accuracy >= 80 else "OK")
        ws.append([sign, st["correct"], st["total"], accuracy, status])

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# 4. CERTIFICATION REPORT — exam history, levels attempted/passed

def generate_certification_report_pdf(user_id: UUID, learner_name: str) -> bytes:
    """Certification report: all exam attempts, levels, scores, pass/fail."""
    exams = certification_store.get_exams_by_user(user_id)
    completed = sorted([e for e in exams if e.status == "completed"],
                       key=lambda e: e.completed_at)

    buffer = io.BytesIO()
    doc, story, styles = _pdf_doc(buffer, "CERTIFICATION REPORT",
                                   "AI-Powered Sign Language Learning Platform", learner_name)
    body = ParagraphStyle("b", parent=styles["Normal"], fontSize=11, spaceAfter=6)

    passed = [e for e in completed if e.passed]
    story.append(Paragraph(f"Exams Completed: <b>{len(completed)}</b> | "
                            f"Passed: <b>{len(passed)}</b>", body))
    story.append(Spacer(1, 0.4*cm))

    table_data = [["Level", "Signs", "Score", "Pass Threshold", "Result", "Date"]]
    for exam in completed:
        accuracy = round(exam.correct_predictions / exam.total_predictions * 100, 1) \
            if exam.total_predictions else 0.0
        table_data.append([
            exam.level.title(),
            str(len(exam.required_signs)),
            f"{exam.score}%",
            f"{exam.pass_threshold}%",
            "PASS" if exam.passed else "FAIL",
            exam.completed_at.strftime("%Y-%m-%d"),
        ])

    t = Table(table_data, colWidths=[3.5*cm, 2*cm, 2.5*cm, 3.5*cm, 2.5*cm, 3*cm])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1565c0")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e0e0e0")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f5f5f5")]),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
    ]))
    story.append(t)

    doc.build(story)
    pdf_bytes = buffer.getvalue()
    buffer.close()
    return pdf_bytes


def generate_certification_report_excel(user_id: UUID) -> bytes:
    """Excel version of the certification report."""
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError("openpyxl not installed")

    exams = certification_store.get_exams_by_user(user_id)
    completed = sorted([e for e in exams if e.status == "completed"],
                       key=lambda e: e.completed_at)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Certification Report"
    _header_row(ws, ["Level", "Signs", "Score (%)", "Pass Threshold (%)", "Result", "Date"])

    for exam in completed:
        ws.append([
            exam.level.title(),
            len(exam.required_signs),
            exam.score,
            exam.pass_threshold,
            "PASS" if exam.passed else "FAIL",
            exam.completed_at.strftime("%Y-%m-%d"),
        ])

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()